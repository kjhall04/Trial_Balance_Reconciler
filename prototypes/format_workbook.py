from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def format_workbook(excel_path: str, money_columns=None, text_wrap_columns=None):
    if money_columns is None:
        money_columns = {
            "Debit", "Credit", "balance_client",
            "balance_import", "balance_diff", "col3"
        }
    if text_wrap_columns is None:
        text_wrap_columns = {"account", "account_client", "account_import", "class"}

    wb = load_workbook(excel_path)

    for ws in wb.worksheets:
        if ws.max_row < 2 or ws.max_column < 1:
            continue

        # Freeze header row
        ws.freeze_panes = "A2"

        # Make header bold and enable filters
        ws.auto_filter.ref = ws.dimensions
        header_font = Font(bold=True)

        # Build a map of column index to header name
        headers = {}
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            headers[col] = str(header_value) if header_value is not None else ""

            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Apply formats and compute widths
        col_widths = {col: 10 for col in range(1, ws.max_column + 1)}

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                header = headers[col]

                # Alignment rules
                if header in text_wrap_columns:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                elif header in money_columns:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                else:
                    # Default align
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="top")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top")

                # Width calculation
                v = cell.value
                if v is None:
                    length = 0
                else:
                    # Keep widths reasonable for long text
                    s = str(v)
                    length = min(len(s), 60)

                if length + 2 > col_widths[col]:
                    col_widths[col] = length + 2

        # Set column widths with a max cap
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = min(max(width, 10), 55)

        # Optional: set row 1 height for readability
        ws.row_dimensions[1].height = 22

    wb.save(excel_path)


# Example usage after you write the Excel file
# out_path is the same path you already used in pd.ExcelWriter