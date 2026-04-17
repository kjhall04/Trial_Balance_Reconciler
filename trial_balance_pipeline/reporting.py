from __future__ import annotations

from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .models import TrialBalanceBuildResult


SHEET_NAME = "Trial Balance"
CONFIDENCE_FONT_COLORS = {
    "high": "2E7D32",
    "medium": "B8860B",
    "low": "C62828",
}
ACCOUNTING_NUMBER_FORMAT = '#,##0.00;(#,##0.00);-'
DETAIL_BALANCE_HEADERS = {"py_balance", "cy_balance"}
REVIEW_BALANCE_HEADERS = {"previous year rep", "current year prelim", "ajes", "final"}


def _is_multi_entity(df: pd.DataFrame) -> bool:
    if "entity" not in df.columns:
        return False
    values = [str(value).strip() for value in df["entity"].tolist() if pd.notna(value) and str(value).strip()]
    return len(set(values)) > 1


def _drop_entity_if_single(df: pd.DataFrame) -> pd.DataFrame:
    if "entity" in df.columns and not _is_multi_entity(df):
        return df.drop(columns=["entity"]).copy()
    return df.copy()


def _set_font_color(cell, color: str) -> None:
    font = copy(cell.font)
    font.color = color
    cell.font = font


def _apply_row_confidence_colors(
    path: str | Path,
    sheet_name: str,
    confidence_levels: list[str],
    target_columns: list[int],
    *,
    start_row: int,
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name]
        for row_number, level in enumerate(confidence_levels, start=start_row):
            color = CONFIDENCE_FONT_COLORS.get(str(level).strip().lower())
            if not color:
                continue
            for column_number in target_columns:
                if column_number > sheet.max_column:
                    continue
                _set_font_color(sheet.cell(row=row_number, column=column_number), color)
    finally:
        workbook.save(path)


def _color_sheet_from_confidence_column(sheet) -> None:
    if sheet.max_row < 2:
        return

    headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value is not None}
    confidence_column = headers.get("confidence_level")
    if confidence_column is None:
        return

    preferred_headers = [
        "assigned_class",
        "assigned_acct_no",
        "assigned_account",
        "class",
        "acct_no",
        "account",
        "confidence_level",
        "decision_note",
    ]
    target_columns = [headers[name] for name in preferred_headers if name in headers]
    if not target_columns:
        target_columns = [confidence_column]

    for row_number in range(2, sheet.max_row + 1):
        level = sheet.cell(row=row_number, column=confidence_column).value
        color = CONFIDENCE_FONT_COLORS.get(str(level).strip().lower())
        if not color:
            continue
        for column_number in target_columns:
            _set_font_color(sheet.cell(row=row_number, column=column_number), color)


def _normalized_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def _balance_columns_for_sheet(sheet) -> set[int]:
    if sheet.title == SHEET_NAME:
        review_headers = {_normalized_header(cell.value): cell.column for cell in sheet[3] if cell.value is not None}
        review_balance_columns = {
            column
            for header, column in review_headers.items()
            if header in REVIEW_BALANCE_HEADERS
        }
        if review_balance_columns:
            return review_balance_columns

        if sheet.max_column >= 2:
            return {sheet.max_column - 1, sheet.max_column}
        return set()

    headers = {_normalized_header(cell.value): cell.column for cell in sheet[1] if cell.value is not None}
    return {
        column
        for header, column in headers.items()
        if header in DETAIL_BALANCE_HEADERS
    }


def _apply_balance_number_format(sheet) -> None:
    balance_columns = _balance_columns_for_sheet(sheet)
    if not balance_columns:
        return

    for row in sheet.iter_rows():
        for cell in row:
            if cell.column not in balance_columns:
                continue
            if isinstance(cell.value, (int, float)):
                cell.number_format = ACCOUNTING_NUMBER_FORMAT


def write_import_workbook(df: pd.DataFrame, out_path: Path, sheet_name: str = SHEET_NAME) -> None:
    export_columns = ["entity", "class", "acct_no", "account", "py_balance", "cy_balance"]
    export_df = df[export_columns].copy()
    if not _is_multi_entity(export_df):
        export_df = export_df.drop(columns=["entity"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    if "confidence_level" in df.columns:
        target_columns = [2, 3, 4] if _is_multi_entity(df) else [1, 2, 3]
        _apply_row_confidence_colors(
            out_path,
            sheet_name,
            df["confidence_level"].fillna("").astype(str).tolist(),
            target_columns,
            start_row=1,
        )


def write_review_workbook(df: pd.DataFrame, out_path: Path, title_text: str, sheet_name: str = SHEET_NAME) -> None:
    use_entity = _is_multi_entity(df)
    headers = (
        ["Entity / Fund", "Leadsheet", "Account Number", "Account Name", "Previous Year Rep", "Current Year Prelim", "AJEs", "Final"]
        if use_entity
        else ["Leadsheet", "Account Number", "Account Name", "Previous Year Rep", "Current Year Prelim", "AJEs", "Final"]
    )
    rows: list[list[object]] = [
        [title_text] + [None] * (len(headers) - 1),
        [None] * len(headers),
        headers,
    ]
    for _, row in df.iterrows():
        if use_entity:
            rows.append(
                [
                    row["entity"],
                    row["class"],
                    row["acct_no"],
                    row["account"],
                    float(row["py_balance"]),
                    float(row["cy_balance"]),
                    0.0,
                    float(row["cy_balance"]),
                ]
            )
        else:
            rows.append(
                [
                    row["class"],
                    row["acct_no"],
                    row["account"],
                    float(row["py_balance"]),
                    float(row["cy_balance"]),
                    0.0,
                    float(row["cy_balance"]),
                ]
            )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    if "confidence_level" in df.columns:
        target_columns = [2, 3, 4] if use_entity else [1, 2, 3]
        _apply_row_confidence_colors(
            out_path,
            sheet_name,
            df["confidence_level"].fillna("").astype(str).tolist(),
            target_columns,
            start_row=4,
        )


def write_details_workbook(result: TrialBalanceBuildResult, out_path: Path) -> None:
    import_ready_df = result.updated_import[
        [
            "entity",
            "class",
            "acct_no",
            "account",
            "py_balance",
            "cy_balance",
            "confidence_level",
            "confidence_score",
            "confidence_reason",
            "decision_note",
            "current_source_reference",
            "matched_prior_reference",
            "source_kind",
        ]
    ].copy()

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("summary", result.summary.copy()),
        ("import_ready_tb", _drop_entity_if_single(import_ready_df)),
        ("audit_trail", _drop_entity_if_single(result.comparison_details)),
        ("current_year_raw", _drop_entity_if_single(result.current_trial_balance)),
        ("prior_year_tb", _drop_entity_if_single(result.prior_year_rows)),
        ("matched_rows", _drop_entity_if_single(result.matched_rows)),
        ("new_rows", _drop_entity_if_single(result.new_rows)),
        ("carryforward_rows", _drop_entity_if_single(result.carryforward_rows)),
        ("renumbered_rows", _drop_entity_if_single(result.renumbered_rows)),
        ("review_queue", _drop_entity_if_single(result.review_queue)),
    ]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets:
            frame.to_excel(writer, sheet_name=sheet_name, index=False)


def format_workbook(path: str | Path, *, freeze_panes: str | None = None, max_width: int = 72) -> None:
    workbook = load_workbook(path)
    try:
        for sheet in workbook.worksheets:
            if freeze_panes:
                sheet.freeze_panes = freeze_panes
            widths: dict[int, int] = {}
            for row in sheet.iter_rows():
                for cell in row:
                    value = "" if cell.value is None else str(cell.value)
                    widths[cell.column] = min(max(widths.get(cell.column, 10), len(value) + 2), max_width)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    if isinstance(cell.value, (int, float)) and cell.column >= max(sheet.max_column - 1, 1):
                        cell.alignment = Alignment(horizontal="right", vertical="top")
            _apply_balance_number_format(sheet)
            _color_sheet_from_confidence_column(sheet)
            for column, width in widths.items():
                sheet.column_dimensions[get_column_letter(column)].width = width
    finally:
        workbook.save(path)


def format_outputs(import_path: str | Path, details_path: str | Path | None, review_path: str | Path | None = None) -> None:
    format_workbook(import_path)
    if details_path:
        format_workbook(details_path, freeze_panes="A2")
    if review_path:
        format_workbook(review_path)
