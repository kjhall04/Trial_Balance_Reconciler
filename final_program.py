"""
Trial Balance MVP Reconciler

What this program does

1) Reads client tb and computes a net CY balance per account as Debit minus Credit
2) Reads tb to import as the upload template format
3) Finds which client accounts are already represented in the import file
4) For client accounts not represented, it tries to avoid double counting by doing this
   a) If an import row has the exact same CY amount, it can rename that import row to the client name
      This only happens when it is safe (unique balance match or strong name similarity)
   b) Otherwise it adds a new row and assigns
      leadsheet based on the most common leadsheet already used in that bucket
      account number as the next available number in that bucket that does not already exist in import
5) Writes an output file that matches the import template exactly
   5 columns, no header row, sheet name Trial Balance

It also optionally writes
details workbook
review style workbook similar to the APR example

Usage

python tb_mvp_reconciler.py

Or with custom paths

python tb_mvp_reconciler.py --client "client tb.xlsx" --import "tb to import.xlsx"
"""

from __future__ import annotations

import argparse
import operator
import re
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from guess_category import guess_category, CATEGORY_RANGES

import pandas as pd


SHEET_NAME = "Trial Balance"

# When multiple import rows share the same balance, only rename if name similarity clears this threshold
NAME_SIM_STRICT = 0.72
SIMILAR_NAME_MIN = 0.80
SIMILAR_NAME_MARGIN = 0.05
ZERO_BALANCE_PLACES = 2


def norm_key(value: str) -> str:
    s = "" if value is None else str(value)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def is_footer_account(value: str) -> bool:
    s = "" if value is None else str(value).strip()
    upper = s.upper()
    if upper == "TOTAL":
        return True
    if re.search(r"\bGMT\b", s):
        return True
    if s.lower().startswith(
        ("monday,", "tuesday,", "wednesday,", "thursday,", "friday,", "saturday,", "sunday,")
    ):
        return True
    return False


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def _to_int_or_none(x):
    try:
        if x is None or str(x).strip() == "":
            return None
        return int(float(x))
    except Exception:
        return None


def balances_differ(a, b, places: int = 2) -> bool:
    left = 0.0 if pd.isna(a) else float(a)
    right = 0.0 if pd.isna(b) else float(b)
    return round(left, places) != round(right, places)


def is_zero_balance(value, places: int = ZERO_BALANCE_PLACES) -> bool:
    amount = 0.0 if pd.isna(value) else float(value)
    return round(amount, places) == 0.0


def norm_match_text(value: str) -> str:
    s = norm_key(value)
    replacements = {
        "&": " and ",
        "a/r": "accounts receivable",
        "a/p": "accounts payable",
        "ltd": "long term debt",
        "n/p": "notes payable",
        "#": " ",
    }
    for source, target in replacements.items():
        s = s.replace(source, target)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def token_overlap_ratio(a: str, b: str) -> float:
    stop_words = {"and", "the", "of", "to", "for", "in", "on", "at", "by", "from", "current"}
    a_tokens = {t for t in norm_match_text(a).split() if t not in stop_words}
    b_tokens = {t for t in norm_match_text(b).split() if t not in stop_words}
    if not a_tokens and not b_tokens:
        return 1.0
    return len(a_tokens & b_tokens) / max(len(a_tokens), len(b_tokens), 1)


def name_match_metrics(client_name: str, import_name: str) -> dict:
    client_match = norm_match_text(client_name)
    import_match = norm_match_text(import_name)
    seq_score = similarity(client_match, import_match)
    overlap_score = token_overlap_ratio(client_name, import_name)
    contained_score = 1.0 if client_match and (
        client_match in import_match or import_match in client_match
    ) else 0.0
    same_category = guess_category(client_name) == guess_category(import_name)

    score = max(seq_score, (seq_score + overlap_score + contained_score) / 3.0)
    if same_category:
        score += 0.04

    return {
        "score": min(score, 1.0),
        "seq_score": seq_score,
        "overlap_score": overlap_score,
        "contained_score": contained_score,
        "same_category": same_category,
    }


def read_client_tb(path: Path, sheet_name: str = SHEET_NAME) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)

    header_row = None
    for i, row in raw.iterrows():
        r = row.astype(str)
        if r.str.contains("Debit", case=False, na=False).any() and r.str.contains("Credit", case=False, na=False).any():
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find client header row containing Debit and Credit")

    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
    df = df.rename(columns={df.columns[0]: "account"})
    df = df[~df["account"].isna()].copy()

    df["account"] = df["account"].astype(str).str.strip()
    df = df[~df["account"].apply(is_footer_account)].copy()

    df["Debit"] = pd.to_numeric(df.get("Debit"), errors="coerce").fillna(0.0)
    df["Credit"] = pd.to_numeric(df.get("Credit"), errors="coerce").fillna(0.0)

    # Net CY balance equals Debit minus Credit, using operator.sub to avoid relying on the minus token
    df["cy_balance"] = df["Debit"].combine(df["Credit"], operator.sub)
    df = df[~df["cy_balance"].apply(is_zero_balance)].copy()

    df["account_key"] = df["account"].map(norm_key)
    df["match_key"] = df["account"].map(norm_match_text)
    df["client_row"] = range(1, len(df) + 1)

    return df[["client_row", "account", "account_key", "match_key", "cy_balance"]]


def read_import_tb(path: Path, sheet_name: str = SHEET_NAME) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name, header=None).iloc[:, :5].copy()
    df.columns = ["class", "acct_no", "account", "py_balance", "cy_balance"]

    df["class"] = df["class"].fillna("").astype(str).str.strip()
    df["account"] = df["account"].fillna("").astype(str).str.strip()

    df["acct_no"] = pd.to_numeric(df["acct_no"], errors="coerce")
    df["py_balance"] = pd.to_numeric(df["py_balance"], errors="coerce").fillna(0.0)
    df["cy_balance"] = pd.to_numeric(df["cy_balance"], errors="coerce").fillna(0.0)

    df["account_key"] = df["account"].map(norm_key)
    df["match_key"] = df["account"].map(norm_match_text)
    df["import_row"] = range(1, len(df) + 1)

    return df


def default_leadsheet_for_bucket(import_df: pd.DataFrame) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for bucket, (lo, hi) in CATEGORY_RANGES.items():
        block = import_df[(import_df["acct_no"] >= lo) & (import_df["acct_no"] <= hi)]
        if len(block) and block["class"].notna().any():
            mapping[bucket] = block["class"].value_counts().idxmax()
        else:
            mapping[bucket] = f"New {bucket.title()}"
    return mapping


def used_numbers(import_df: pd.DataFrame) -> set[int]:
    return set(import_df["acct_no"].dropna().astype(int).tolist())


def next_available_in_bucket(import_df: pd.DataFrame, bucket: str) -> int:
    lo, hi = CATEGORY_RANGES[bucket]
    used = used_numbers(import_df)
    for n in range(lo, hi + 1):
        if n not in used:
            return n
    return hi


@dataclass
class MVPResult:
    updated_import: pd.DataFrame
    summary: pd.DataFrame
    renamed_existing_rows: pd.DataFrame
    changed_existing_rows: pd.DataFrame
    new_rows_added: pd.DataFrame
    removed_import_rows: pd.DataFrame


def reconcile_mvp(client_df: pd.DataFrame, import_df: pd.DataFrame) -> MVPResult:
    client_total = round(float(client_df["cy_balance"].sum()), ZERO_BALANCE_PLACES)
    if client_total != 0.0:
        raise ValueError(f"Client trial balance does not net to zero. Current total is {client_total:.2f}.")

    used_client_rows: set[int] = set()
    used_import_rows: set[int] = set()
    existing_rows: List[dict] = []
    rename_pairs: List[dict] = []
    changed_existing_rows_records: List[dict] = []

    def add_existing_match(client_row: pd.Series, import_row: pd.Series, match_rule: str, name_similarity: float | None = None) -> None:
        updated_row = import_row.to_dict()
        updated_row["account"] = client_row["account"]
        updated_row["account_key"] = norm_key(client_row["account"])
        updated_row["match_key"] = norm_match_text(client_row["account"])
        updated_row["cy_balance"] = float(client_row["cy_balance"])
        existing_rows.append(updated_row)

        old_cy = float(import_row["cy_balance"])
        new_cy = float(client_row["cy_balance"])
        if balances_differ(old_cy, new_cy):
            changed_existing_rows_records.append(
                {
                    "import_row": int(import_row["import_row"]),
                    "acct_no": float(import_row["acct_no"]) if pd.notna(import_row["acct_no"]) else None,
                    "account": client_row["account"],
                    "old_cy_balance": old_cy,
                    "new_cy_balance": new_cy,
                    "match_rule": match_rule,
                }
            )

        import_account = "" if pd.isna(import_row["account"]) else str(import_row["account"]).strip()
        client_account = "" if pd.isna(client_row["account"]) else str(client_row["account"]).strip()
        if import_account != client_account:
            rename_pairs.append(
                {
                    "client_row": int(client_row["client_row"]),
                    "client_account": client_account,
                    "client_cy": float(client_row["cy_balance"]),
                    "import_row": int(import_row["import_row"]),
                    "import_acct_no": float(import_row["acct_no"]) if pd.notna(import_row["acct_no"]) else None,
                    "import_account": import_account,
                    "import_cy": float(import_row["cy_balance"]),
                    "name_similarity": round(float(name_similarity or 0.0), 3),
                    "match_rule": match_rule,
                }
            )

        used_client_rows.add(int(client_row["client_row"]))
        used_import_rows.add(int(import_row["import_row"]))

    # Step 1 exact account name matches
    exact_candidates = client_df.merge(import_df, on="account_key", how="inner", suffixes=("_client", "_import"))
    exact_candidates = exact_candidates.sort_values(["client_row", "import_row"], kind="mergesort")
    for _, row in exact_candidates.iterrows():
        client_row_no = int(row["client_row"])
        import_row_no = int(row["import_row"])
        if client_row_no in used_client_rows or import_row_no in used_import_rows:
            continue
        client_row = client_df.loc[client_df["client_row"] == client_row_no].iloc[0]
        import_row = import_df.loc[import_df["import_row"] == import_row_no].iloc[0]
        add_existing_match(client_row, import_row, "exact_name", name_similarity=1.0)

    # Step 2 strong similar-name matches, even when the balance changed
    client_remaining = client_df[~client_df["client_row"].isin(used_client_rows)].copy()
    import_remaining = import_df[~import_df["import_row"].isin(used_import_rows)].copy()

    similar_candidates: List[dict] = []
    for _, crow in client_remaining.iterrows():
        scored: List[dict] = []
        for _, irow in import_remaining.iterrows():
            metrics = name_match_metrics(crow["account"], irow["account"])
            strong_name_match = metrics["score"] >= SIMILAR_NAME_MIN and (
                metrics["same_category"] or metrics["seq_score"] >= 0.86 or metrics["contained_score"] == 1.0
            )
            if not strong_name_match:
                continue
            scored.append(
                {
                    "client_row": int(crow["client_row"]),
                    "import_row": int(irow["import_row"]),
                    "score": float(metrics["score"]),
                }
            )

        if not scored:
            continue

        scored.sort(key=lambda x: x["score"], reverse=True)
        best = scored[0]
        second_score = scored[1]["score"] if len(scored) > 1 else -1.0
        if best["score"] >= 0.94 or best["score"] - second_score >= SIMILAR_NAME_MARGIN:
            similar_candidates.append(best)

    similar_candidates.sort(key=lambda x: x["score"], reverse=True)
    for candidate in similar_candidates:
        client_row_no = candidate["client_row"]
        import_row_no = candidate["import_row"]
        if client_row_no in used_client_rows or import_row_no in used_import_rows:
            continue
        client_row = client_df.loc[client_df["client_row"] == client_row_no].iloc[0]
        import_row = import_df.loc[import_df["import_row"] == import_row_no].iloc[0]
        add_existing_match(client_row, import_row, "similar_name", name_similarity=candidate["score"])

    # Step 3 same-balance fallback for names that changed materially but still appear to be the same account
    client_remaining = client_df[~client_df["client_row"].isin(used_client_rows)].copy()
    import_remaining = import_df[~import_df["import_row"].isin(used_import_rows)].copy()

    bal_index: Dict[float, List[int]] = defaultdict(list)
    for _, row in import_remaining.iterrows():
        bal_index[round(float(row["cy_balance"]), ZERO_BALANCE_PLACES)].append(int(row["import_row"]))

    for _, crow in client_remaining.iterrows():
        bal = round(float(crow["cy_balance"]), ZERO_BALANCE_PLACES)
        candidate_import_rows = [row_no for row_no in bal_index.get(bal, []) if row_no not in used_import_rows]
        if not candidate_import_rows:
            continue

        if len(candidate_import_rows) == 1:
            import_row_no = candidate_import_rows[0]
            import_row = import_df.loc[import_df["import_row"] == import_row_no].iloc[0]
            name_score = similarity(crow["account_key"], import_row["account_key"])
            add_existing_match(crow, import_row, "unique_balance", name_similarity=name_score)
            continue

        best_import_row, best_score = None, -1.0
        for import_row_no in candidate_import_rows:
            import_row = import_df.loc[import_df["import_row"] == import_row_no].iloc[0]
            metrics = name_match_metrics(crow["account"], import_row["account"])
            if metrics["score"] > best_score:
                best_import_row, best_score = import_row_no, metrics["score"]

        if best_import_row is not None and best_score >= NAME_SIM_STRICT:
            import_row = import_df.loc[import_df["import_row"] == best_import_row].iloc[0]
            add_existing_match(crow, import_row, "balance_plus_name_strict", name_similarity=best_score)

    renamed_existing_rows = pd.DataFrame(
        rename_pairs,
        columns=[
            "client_row",
            "client_account",
            "client_cy",
            "import_row",
            "import_acct_no",
            "import_account",
            "import_cy",
            "name_similarity",
            "match_rule",
        ],
    )
    changed_existing_rows = pd.DataFrame(
        changed_existing_rows_records,
        columns=[
            "import_row",
            "acct_no",
            "account",
            "old_cy_balance",
            "new_cy_balance",
            "match_rule",
        ],
    )

    removed_import_records: List[dict] = []
    for _, row in import_df[~import_df["import_row"].isin(used_import_rows)].iterrows():
        old_cy = float(row["cy_balance"])
        removal_reason = (
            "Removed because the import CY balance is zero."
            if is_zero_balance(old_cy)
            else "Removed because no matching nonzero client account was found."
        )
        removed_import_records.append(
            {
                "import_row": int(row["import_row"]),
                "acct_no": float(row["acct_no"]) if pd.notna(row["acct_no"]) else None,
                "class": row["class"],
                "account": row["account"],
                "py_balance": float(row["py_balance"]),
                "cy_balance": old_cy,
                "removal_reason": removal_reason,
            }
        )

    removed_import_rows = pd.DataFrame(
        removed_import_records,
        columns=[
            "import_row",
            "acct_no",
            "class",
            "account",
            "py_balance",
            "cy_balance",
            "removal_reason",
        ],
    )

    client_still_missing = client_df[~client_df["client_row"].isin(used_client_rows)].copy()

    # Step 4 add truly new rows using next available number in bucket
    leadsheet_map = default_leadsheet_for_bucket(import_df)
    used_numbers_set = set(import_df["acct_no"].dropna().astype(int).tolist())
    existing_import_row_max = int(import_df["import_row"].max()) if len(import_df) else 0

    def next_available_in_bucket_used(used_set: set[int], bucket: str) -> int:
        lo, hi = CATEGORY_RANGES[bucket]
        for n in range(lo, hi + 1):
            if n not in used_set:
                used_set.add(n)
                return n
        used_set.add(hi)
        return hi

    new_rows: List[dict] = []
    for _, row in client_still_missing.iterrows():
        name = row["account"]
        bucket = guess_category(name)
        acct_no = next_available_in_bucket_used(used_numbers_set, bucket)
        leadsheet = leadsheet_map.get(bucket, f"New {bucket.title()}")
        new_rows.append(
            {
                "class": leadsheet,
                "acct_no": int(acct_no),
                "account": name,
                "py_balance": 0.0,
                "cy_balance": float(row["cy_balance"]),
                "account_key": norm_key(name),
                "match_key": norm_match_text(name),
                "import_row": existing_import_row_max + 1 + len(new_rows),
                "source_client_row": int(row["client_row"]),
                "bucket": bucket,
            }
        )

    new_rows_added = pd.DataFrame(
        new_rows,
        columns=[
            "class",
            "acct_no",
            "account",
            "py_balance",
            "cy_balance",
            "account_key",
            "match_key",
            "import_row",
            "source_client_row",
            "bucket",
        ],
    )

    updated_existing = pd.DataFrame(existing_rows, columns=import_df.columns)
    if len(new_rows_added):
        updated = pd.concat(
            [updated_existing, new_rows_added.drop(columns=["source_client_row", "bucket"], errors="ignore")],
            ignore_index=True,
        )
    else:
        updated = updated_existing

    # Safety checks
    bad = updated[updated["account"].isna() | (updated["account"].astype(str).str.strip() == "")]
    if len(bad):
        raise ValueError("Found blank account rows in updated output. Placeholder rows still exist or bad data was appended.")

    updated = updated[~updated["cy_balance"].apply(is_zero_balance)].copy()
    updated_sorted = updated.sort_values(["acct_no", "account"], kind="mergesort").reset_index(drop=True)

    updated_total = round(float(updated_sorted["cy_balance"].sum()), ZERO_BALANCE_PLACES)
    if updated_total != 0.0:
        raise ValueError(f"Updated import does not net to zero. Current total is {updated_total:.2f}.")

    summary = pd.DataFrame(
        [
            {"metric": "client cy total", "value": round(float(client_df["cy_balance"].sum()), ZERO_BALANCE_PLACES)},
            {"metric": "import cy total before", "value": round(float(import_df["cy_balance"].sum()), ZERO_BALANCE_PLACES)},
            {"metric": "import cy total after", "value": round(float(updated_sorted["cy_balance"].sum()), ZERO_BALANCE_PLACES)},
            {"metric": "rename count", "value": int(len(renamed_existing_rows))},
            {"metric": "existing balance updates", "value": int(len(changed_existing_rows))},
            {"metric": "new account count", "value": int(len(new_rows_added))},
            {"metric": "deleted import row count", "value": int(len(removed_import_rows))},
            {
                "metric": "deleted zero-balance import rows",
                "value": int(len(removed_import_rows[removed_import_rows["cy_balance"].apply(is_zero_balance)])),
            },
            {
                "metric": "deleted nonzero import rows",
                "value": int(len(removed_import_rows[~removed_import_rows["cy_balance"].apply(is_zero_balance)])),
            },
            {"metric": "output row count", "value": int(len(updated_sorted))},
        ]
    )

    return MVPResult(
        updated_import=updated_sorted,
        summary=summary,
        renamed_existing_rows=renamed_existing_rows,
        changed_existing_rows=changed_existing_rows,
        new_rows_added=new_rows_added,
        removed_import_rows=removed_import_rows,
    )


def write_import_format(df: pd.DataFrame, out_path: Path, sheet_name: str = SHEET_NAME) -> None:
    export_df = df[["class", "acct_no", "account", "py_balance", "cy_balance"]].copy()

    export_df["class"] = export_df["class"].astype(str).str.strip()
    export_df["account"] = export_df["account"].astype(str).str.strip()

    export_df["acct_no"] = pd.to_numeric(export_df["acct_no"], errors="coerce")
    export_df["py_balance"] = pd.to_numeric(export_df["py_balance"], errors="coerce").fillna(0.0)
    export_df["cy_balance"] = pd.to_numeric(export_df["cy_balance"], errors="coerce").fillna(0.0)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # No headers, same as your upload file
        export_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)


def _clean_money_value(value):
    if pd.isna(value):
        return None
    amount = float(value)
    if is_zero_balance(amount):
        return 0.0
    return round(amount, ZERO_BALANCE_PLACES)


def _clean_money_columns(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if col in out.columns:
            out[col] = out[col].apply(_clean_money_value)
    return out


def _clean_int_columns(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").astype("Int64")
    return out


def _friendly_match_rule(rule: str) -> str:
    labels = {
        "exact_name": "Exact name match",
        "similar_name": "Strong similar-name match",
        "unique_balance": "Unique balance match",
        "balance_plus_name_strict": "Balance + strong name match",
    }
    return labels.get(rule, str(rule).replace("_", " ").title())


def _build_summary_sheet(summary_df: pd.DataFrame) -> pd.DataFrame:
    labels = {
        "client cy total": ("Client CY total", "Client trial balance total after zero-balance lines are excluded."),
        "import cy total before": ("Import CY total before update", "Original import workbook total before any changes."),
        "import cy total after": ("Import CY total after update", "Updated import workbook total after all changes."),
        "rename count": ("Renamed rows", "Existing import rows where the account name changed."),
        "existing balance updates": ("Balance updates", "Existing import rows where the CY amount changed."),
        "new account count": ("New rows added", "Client accounts added as brand-new import rows."),
        "deleted import row count": ("Rows removed", "Import rows removed from the updated import file."),
        "deleted zero-balance import rows": ("Zero-balance rows removed", "Removed import rows whose CY balance was effectively 0.00."),
        "deleted nonzero import rows": ("Nonzero rows removed", "Removed import rows that did not match a nonzero client account."),
        "output row count": ("Final output row count", "Row count in the updated import workbook."),
    }

    display_rows: List[dict] = []
    for _, row in summary_df.iterrows():
        metric = str(row["metric"])
        value = row["value"]
        title, meaning = labels.get(metric, (metric.replace("_", " ").title(), ""))
        if "total" in metric:
            value_display = f"{_clean_money_value(value):,.2f}"
        else:
            value_display = str(int(float(value)))
        display_rows.append(
            {
                "Metric": title,
                "Meaning": meaning,
                "Value": value_display,
            }
        )
    return pd.DataFrame(display_rows, columns=["Metric", "Meaning", "Value"])


def _build_renamed_sheet(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "Import Row",
                "Account No",
                "Old Account Name",
                "Old CY Balance",
                "New CY Balance",
                "New Account Name",
                "Match Method",
                "Name Similarity",
            ]
        )

    display = pd.DataFrame(
        {
            "Import Row": df["import_row"],
            "Account No": df["import_acct_no"],
            "Old Account Name": df["import_account"],
            "Old CY Balance": df["import_cy"],
            "New CY Balance": df["client_cy"],
            "New Account Name": df["client_account"],
            "Match Method": df["match_rule"].map(_friendly_match_rule),
            "Name Similarity": pd.to_numeric(df["name_similarity"], errors="coerce").map(
                lambda x: "" if pd.isna(x) else f"{float(x):.3f}"
            ),
        }
    )
    display = _clean_int_columns(display, ["Import Row", "Account No"])
    display = _clean_money_columns(display, ["Old CY Balance", "New CY Balance"])
    return display


def _build_changed_sheet(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "Import Row",
                "Account No",
                "Account Name",
                "Old CY Balance",
                "New CY Balance",
                "Match Method",
                "Update Summary",
                "Difference",
            ]
        )

    old_clean = df["old_cy_balance"].apply(_clean_money_value)
    new_clean = df["new_cy_balance"].apply(_clean_money_value)
    diff = pd.to_numeric(new_clean, errors="coerce").fillna(0.0) - pd.to_numeric(old_clean, errors="coerce").fillna(0.0)
    display = pd.DataFrame(
        {
            "Import Row": df["import_row"],
            "Account No": df["acct_no"],
            "Account Name": df["account"],
            "Old CY Balance": old_clean,
            "New CY Balance": new_clean,
            "Match Method": df["match_rule"].map(_friendly_match_rule),
            "Update Summary": "CY balance updated from the client TB.",
            "Difference": diff,
        }
    )
    display = _clean_int_columns(display, ["Import Row", "Account No"])
    display = _clean_money_columns(display, ["Old CY Balance", "New CY Balance", "Difference"])
    return display


def _build_new_rows_sheet(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "Client Row",
                "Account No",
                "Account Name",
                "PY Balance",
                "CY Balance",
                "Leadsheet",
                "Category Bucket",
            ]
        )

    display = pd.DataFrame(
        {
            "Client Row": df["source_client_row"],
            "Account No": df["acct_no"],
            "Account Name": df["account"],
            "PY Balance": df["py_balance"],
            "CY Balance": df["cy_balance"],
            "Leadsheet": df["class"],
            "Category Bucket": df["bucket"].astype(str).str.replace("_", " ").str.title(),
        }
    )
    display = _clean_int_columns(display, ["Client Row", "Account No"])
    display = _clean_money_columns(display, ["PY Balance", "CY Balance"])
    return display


def _build_removed_rows_sheet(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "Import Row",
                "Account No",
                "Account Name",
                "PY Balance",
                "CY Balance",
                "Leadsheet",
                "Removal Reason",
            ]
        )

    display = pd.DataFrame(
        {
            "Import Row": df["import_row"],
            "Account No": df["acct_no"],
            "Account Name": df["account"],
            "PY Balance": df["py_balance"],
            "CY Balance": df["cy_balance"],
            "Leadsheet": df["class"],
            "Removal Reason": df["removal_reason"],
        }
    )
    display = _clean_int_columns(display, ["Import Row", "Account No"])
    display = _clean_money_columns(display, ["PY Balance", "CY Balance"])
    return display


def write_details_workbook(result: MVPResult, out_path: Path) -> None:
    summary_sheet = _build_summary_sheet(result.summary)
    renamed_sheet = _build_renamed_sheet(result.renamed_existing_rows)
    changed_sheet = _build_changed_sheet(result.changed_existing_rows)
    new_rows_sheet = _build_new_rows_sheet(result.new_rows_added)
    removed_rows_sheet = _build_removed_rows_sheet(result.removed_import_rows)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_sheet.to_excel(writer, index=False, sheet_name="summary")
        renamed_sheet.to_excel(writer, index=False, sheet_name="renamed_existing_rows")
        changed_sheet.to_excel(writer, index=False, sheet_name="changed_existing_rows")
        new_rows_sheet.to_excel(writer, index=False, sheet_name="new_rows_added")
        removed_rows_sheet.to_excel(writer, index=False, sheet_name="deleted_import_rows")


def write_review_style(df: pd.DataFrame, out_path: Path, title_text: str, sheet_name: str = SHEET_NAME) -> None:
    rows: List[List[object]] = []
    rows.append([title_text, None, None, None, None, None, None])
    rows.append([None, None, None, None, "Output column", None, None])
    rows.append(["Leadsheet", "Account Number", "Account Name", "Previous Year Rep", "Current Year Prelim", "AJEs", "Final"])

    for _, r in df.iterrows():
        rows.append(
            [
                r["class"],
                int(r["acct_no"]) if pd.notna(r["acct_no"]) else None,
                r["account"],
                float(r["py_balance"]),
                float(r["cy_balance"]),
                0.0,
                float(r["cy_balance"]),
            ]
        )

    review_df = pd.DataFrame(rows)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        review_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--client", type=str, default="client tb.xlsx")
    p.add_argument("--import_file", type=str, default="tb to import.xlsx")
    p.add_argument("--out_import", type=str, default="tb_to_import_updated.xlsx")
    p.add_argument("--out_details", type=str, default="tb_mvp_details.xlsx")
    p.add_argument("--out_review", type=str, default="tb_review_style.xlsx")
    p.add_argument("--write_review", action="store_true")
    p.add_argument("--review_title", type=str, default="Review Trial Balance")
    return p.parse_args()

def format_workbook(
    path: str | Path,
    wrap_cols: set[int] | None = None,
    money_cols: set[int] | None = None,
    freeze_panes: str | None = None,
    has_header: bool = False,
    max_col_width: int = 60,
    red_acct_numbers : set[int] | None = None,
    red_value_rows: set[Tuple[int | None, str]] | None = None,
    red_account_rows: set[Tuple[int | None, str]] | None = None,
):
    path = Path(path)
    wrap_cols = wrap_cols or set()
    money_cols = money_cols or set()
    red_acct_numbers = red_acct_numbers or set()
    red_value_rows = red_value_rows or set()
    red_account_rows = red_account_rows or set()

    wb = load_workbook(path)

    red_font = Font(color="FF0000")

    for ws in wb.worksheets:
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        if freeze_panes:
            ws.freeze_panes = freeze_panes

        col_widths = {c: 10 for c in range(1, max_col + 1)}

        for r in range(1, max_row + 1):
            # Determine if this row should be red based on acct no in column B
            acct_val = _to_int_or_none(ws.cell(row=r, column=2).value)
            account_val = ws.cell(row=r, column=3).value
            row_key = (acct_val, "" if account_val is None else str(account_val).strip())
            make_row_red = acct_val in red_acct_numbers
            make_value_red = row_key in red_value_rows
            make_account_red = row_key in red_account_rows

            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value

                if c in wrap_cols:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                elif c in money_cols:
                    # Parentheses for negatives, and blanks for zero values
                    cell.number_format = '#,##0.00;(#,##0.00);""'
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                else:
                    if isinstance(val, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="top")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top")

                if make_row_red or (make_account_red and c == 3) or (make_value_red and c == 5):
                    cell.font = red_font

                if val is None:
                    length = 0
                else:
                    length = min(len(str(val)), 70)

                if length + 2 > col_widths[c]:
                    col_widths[c] = length + 2

        for c, width in col_widths.items():
            ws.column_dimensions[get_column_letter(c)].width = min(max(width, 10), max_col_width)

    wb.save(path)


def format_outputs(
    out_import_path: str | Path,
    out_details_path: str | Path | None,
    out_review_path: str | Path | None = None,
    new_accts=None,
    changed_balance_rows: pd.DataFrame | None = None,
    renamed_rows: pd.DataFrame | None = None,
):
    changed_value_keys: set[Tuple[int | None, str]] = set()
    if changed_balance_rows is not None and len(changed_balance_rows):
        for _, row in changed_balance_rows.iterrows():
            changed_value_keys.add(
                (
                    _to_int_or_none(row.get("acct_no")),
                    "" if row.get("account") is None else str(row.get("account")).strip(),
                )
            )

    renamed_row_keys: set[Tuple[int | None, str]] = set()
    if renamed_rows is not None and len(renamed_rows):
        for _, row in renamed_rows.iterrows():
            renamed_row_keys.add(
                (
                    _to_int_or_none(row.get("import_acct_no")),
                    "" if row.get("client_account") is None else str(row.get("client_account")).strip(),
                )
            )

    # Import ready file: no header, columns:
    # A class, B acct_no, C account, D py, E cy
    format_workbook(
        out_import_path,
        wrap_cols={3},
        money_cols={4, 5},
        freeze_panes=None,
        has_header=False,
        red_acct_numbers=set(new_accts or []),
        red_value_rows=changed_value_keys,
        red_account_rows=renamed_row_keys,
    )

    # Details workbook: has headers in each sheet
    if out_details_path:
        format_workbook(
            out_details_path,
            wrap_cols={2, 3, 6, 7},
            money_cols={3, 4, 5, 6, 7, 8, 9},
            freeze_panes="A2",
            has_header=True,
        )

    # Optional review layout
    if out_review_path:
        format_workbook(
            out_review_path,
            wrap_cols={3},
            money_cols={4, 5, 6, 7},
            freeze_panes=None,
            has_header=False,
        )

def main() -> None:
    args = parse_args()

    client_path = Path(args.client)
    import_path = Path(args.import_file)

    client_df = read_client_tb(client_path)
    import_df = read_import_tb(import_path)

    result = reconcile_mvp(client_df, import_df)

    write_import_format(result.updated_import, Path(args.out_import))
    write_details_workbook(result, Path(args.out_details))

    if args.write_review:
        write_review_style(result.updated_import, Path(args.out_review), args.review_title)

    new_acct_set = set()
    if result.new_rows_added is not None and len(result.new_rows_added) and "acct_no" in result.new_rows_added.columns:
        new_acct_set = set(result.new_rows_added["acct_no"].dropna().astype(int).tolist())

    format_outputs(
        Path(args.out_import),
        Path(args.out_details),
        Path(args.out_review) if args.write_review else None,
        new_accts=new_acct_set,
        changed_balance_rows=result.changed_existing_rows,
        renamed_rows=result.renamed_existing_rows,
    )

    print(result.summary.to_string(index=False))


if __name__ == "__main__":
    main()
